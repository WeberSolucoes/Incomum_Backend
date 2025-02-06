from concurrent.futures import ThreadPoolExecutor
import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook, load_workbook
import io
import os
from rest_framework.response import Response
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from incomum.serializers.vendedorSerializer import VendedorSerializer
from incomum.serializers import agenciaSerializer
from incomum.serializers.relatorioSerializer import RelatorioSerializer
from incomum.serializers.lojaSerializer import LojaSerializer
from incomum.serializers.areaComercialSerializer import AreaComercialSerializer
from datetime import datetime
from django.db import connection
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.db.models import Subquery, OuterRef, Sum
from ..models import *
from ..serializers.relatorioSerializer import *
from django.db.models.functions import Round
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def total_byfilter(request) -> Response:
    data_inicio = request.query_params.get("dataInicio")
    data_fim = request.query_params.get("dataFim")


    if data_inicio is None or data_fim is None:
        return Response(
            {"message": "Os parâmetros data inicial e data final são obrigatórios."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        data_inicio = datetime.datetime.strptime(data_inicio, "%d-%m-%Y")
        data_fim = datetime.datetime.strptime(data_fim, "%d-%m-%Y")
        if data_fim < data_inicio:
            return Response(
                {"message": "A data final deve ser maior que a data inicial."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except ValueError:
        return Response(
            {"message": "Data inválida. Use o formato DD-MM-YYYY."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    unidades = request.query_params.getlist("unidade")
    areaComerciais = request.query_params.getlist("areaComercial")
    agencias = request.query_params.getlist("agencia")
    vendedores = request.query_params.getlist("vendedor")

    relatorios = (
        Relatorio.objects.all()
    )
    if len(unidades) > 0:
        relatorios = relatorios.all()

    if len(areaComerciais) > 0:
        relatorios = relatorios.all()

    if len(agencias) > 0:
        relatorios = relatorios.filter(age_codigo__in=agencias)

    if len(vendedores) > 0:
        relatorios = relatorios.filter(ven_codigo__in=vendedores)
    totais = relatorios.aggregate(
        total_valorliquido=Sum("fim_valorliquido"),
        total_valorinc=Sum("fim_valorinc"),
        total_valorincajustado=Sum("fim_valorincajustado"),
    )
    if totais["total_valorliquido"] is None:
        totais["total_valorliquido"] = 0
    if totais["total_valorinc"] is None:
        totais["total_valorinc"] = 0
    if totais["total_valorincajustado"] is None:
        totais["total_valorincajustado"] = 0
    totais["total_valorliquido"] = locale.currency(
        totais["total_valorliquido"], grouping=True
    )
    totais["total_valorinc"] = locale.currency(totais["total_valorinc"], grouping=True)
    totais["total_valorincajustado"] = locale.currency(
        totais["total_valorincajustado"], grouping=True
    )
    return Response(totais, status=status.HTTP_200_OK)

def list_all_byfilter(request):
    user_id = request.user.id
    data_consulta = request.GET.get('dataInicio')
    data_consulta_final = request.GET.get('dataFim')
    unidade_selecionada = request.GET.get('unidades')
    areas_selecionadas = request.GET.getlist('areasComerciais')
    agencia_selecionada = request.GET.getlist('agencias')
    vendedor_selecionada = request.GET.getlist('vendedores')

    # Consultando as áreas do usuário
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ac.aco_codigo
            FROM usuariocomercial ac
            INNER JOIN auth_user uac ON uac.id = ac.usr_codigo
            WHERE uac.id = %s
        """, [user_id])
        user_areas = [row[0] for row in cursor.fetchall()]

    # Validando as datas
    try:
        data_consulta_dt = datetime.strptime(data_consulta, "%Y-%m-%d")
        data_consulta_final_dt = datetime.strptime(data_consulta_final, "%Y-%m-%d")
    except ValueError:
        return Response(
            {"message": "Datas inválidas. Use o formato YYYY-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST
        )

    query = """
        SELECT fim_tipo, tur_numerovenda, tur_codigo, fim_valorliquido, fim_data, fim_markup, fim_valorinc, fim_valorincajustado, aco_descricao, age_descricao, ven_descricao, fat_valorvendabruta
        FROM faturamentosimplificado 
        WHERE fim_data BETWEEN %s AND %s 
    """
    params = [data_consulta_dt, data_consulta_final_dt]

    # Inclua a condição aco_codigo somente se user_areas não estiver vazio
    if user_areas:
        query += " AND aco_codigo IN %s"
        params.append(tuple(user_areas))

    if unidade_selecionada:
        query += " AND loj_codigo = %s"
        params.append(unidade_selecionada)

    if areas_selecionadas:
        query += " AND aco_codigo IN %s"
        params.append(tuple(areas_selecionadas))

    if agencia_selecionada:
        query += " AND age_codigo IN %s"
        params.append(tuple(agencia_selecionada))

    if vendedor_selecionada:
        query += " AND ven_codigo IN %s"
        params.append(tuple(vendedor_selecionada))

    query += " ORDER BY fim_data"

    # Executa a consulta e formata os resultados
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        resultados = cursor.fetchall()

    # Formatando os resultados em dicionários
    resultados_formatados = [{
        'fim_tipo': resultado[0],
        'tur_numerovenda': resultado[1],
        'tur_codigo': resultado[2],
        'fim_valorliquido': resultado[3],
        'fim_data': resultado[4],
        'fim_markup': resultado[5],
        'fim_valorinc': resultado[6],
        'fim_valorincajustado': resultado[7],
        'aco_descricao': resultado[8],
        'age_descricao': resultado[9],
        'ven_descricao': resultado[10],
        'fat_valorvendabruta': resultado[11],
    } for resultado in resultados]

    # Paginação
    paginator = Paginator(resultados_formatados, 100000)  # 100 resultados por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculando as somas após aplicar os filtros
    soma_totais = {
        'total_valorinc': sum(item['fim_valorinc'] for item in resultados_formatados if item['fim_valorinc'] is not None),
        'total_valorincajustado': sum(item['fim_valorincajustado'] for item in resultados_formatados if item['fim_valorincajustado'] is not None),
        'total_valorliquido': sum(item['fim_valorliquido'] for item in resultados_formatados if item['fim_valorliquido'] is not None),
    }

    return Response({
        'resultados': [RelatorioSerializer(resultado).data for resultado in page_obj],
        'num_paginas': paginator.num_pages,
        'pagina_atual': page_obj.number,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'soma_totais': soma_totais,
    }, status=status.HTTP_200_OK)



def list_all_lojas_byfilter(request):
    # Filtrar as lojas conforme necessário, por exemplo:
    # lojas = Loja.objects.filter(user=request.user)
    lojas = Loja.objects.all().order_by('loj_descricao')  # Obtém todas as lojas

    # Serializa as lojas
    serializer = LojaSerializer(lojas, many=True)

    # Retorna a resposta com os dados serializados
    return Response(serializer.data)

def list_all_areas_byfilter(request) -> Response:
    user_id = request.user.id

    # Verificar áreas comerciais associadas ao usuário
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ac.aco_codigo, a.aco_descricao
            FROM usuariocomercial ac
            INNER JOIN areacomercial a ON a.aco_codigo = ac.aco_codigo
            WHERE ac.usr_codigo = %s
        """, [user_id])
        user_areas = cursor.fetchall()

    if request.method == 'GET':
        unidade_selecionada = request.GET.get('unidade')

        if user_areas:
            # Se o usuário tem áreas associadas, mostrar apenas essas áreas
            associacoes = [
                {'aco_codigo': row[0], 'aco_descricao': row[1]}
                for row in user_areas
            ]
        else:
            # Se o usuário não tem áreas associadas, realizar a consulta com base na unidade
            with connection.cursor() as cursor:
                if unidade_selecionada:
                    cursor.execute("""
                        SELECT lc.aco_codigo, a.aco_descricao
                        FROM lojacomercial lc
                        INNER JOIN areacomercial a ON lc.aco_codigo = a.aco_codigo
                        WHERE lc.loj_codigo = %s ORDER BY aco_descricao
                    """, [unidade_selecionada])
                else:
                    cursor.execute("""
                        SELECT a.aco_codigo, a.aco_descricao
                        FROM areacomercial a ORDER BY aco_descricao
                    """)

                resultados = cursor.fetchall()

            # Formatar os resultados como uma lista de dicionários
            associacoes = [
                {'aco_codigo': row[0], 'aco_descricao': row[1]}
                for row in resultados
            ]

        return Response({'associacoes': associacoes})



def list_all_vendedores_byfilter(request):
    user_id = request.user.id  # ID do usuário logado
    unidade_id = request.GET.get('unidade')  # Unidade selecionada, se houver

    vendedores = []

    # Obtém os códigos de áreas comerciais associadas ao usuário logado
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ac.aco_codigo
            FROM usuariocomercial ac
            INNER JOIN auth_user uac ON uac.id = ac.usr_codigo
            WHERE uac.id = %s
        """, [user_id])
        user_areas = [row[0] for row in cursor.fetchall()]

    # Se uma unidade for selecionada, buscar vendedores por unidade
    if unidade_id:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT v.ven_codigo, v.ven_descricao
                FROM faturamentosimplificado f
                JOIN vendedor v ON v.ven_codigo = f.ven_codigo
                WHERE f.loj_codigo = %s
                ORDER BY v.ven_descricao
            """, [unidade_id])
            vendedores = cursor.fetchall()

    # Caso contrário, retorna todos os vendedores
    else:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT v.ven_codigo, v.ven_descricao
                FROM vendedor v
                ORDER BY v.ven_descricao
            """)
            vendedores = cursor.fetchall()

    vendedores_formatados = [{'ven_codigo': ven[0], 'ven_descricao': ven[1]} for ven in vendedores]

    return Response({'vendedores': vendedores_formatados})


def list_all_agencias_byfilter(request) -> Response:
    user_id = request.user.id  # Obtém o ID do usuário logado

    # Obter o parâmetro area_comercial da consulta
    area_comercial = request.GET.getlist('area_comercial[]')  # Isso captura todos os valores em uma lista

    # Obter o aco_codigo associado ao usuário
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ac.aco_codigo
            FROM usuariocomercial ac
            INNER JOIN auth_user uac ON uac.id = ac.usr_codigo
            WHERE uac.id = %s
        """, [user_id])
        user_aco_codigos = [row[0] for row in cursor.fetchall()]

    with connection.cursor() as cursor:
        if area_comercial:
            # Ajusta o formato da área comercial para uma tupla
            area_comercial_tuple = tuple(area_comercial) if len(area_comercial) > 1 else (area_comercial[0],)
            print(f'Filtro por área comercial: {area_comercial_tuple}')
            cursor.execute("""
                SELECT age_codigo, age_descricao
                FROM agencia
                WHERE aco_codigo IN %s
            """, [area_comercial_tuple])
        else:
            # Se nenhuma área comercial foi selecionada, verificar user_aco_codigos
            if user_aco_codigos:
                cursor.execute("""
                    SELECT age_codigo, age_descricao
                    FROM agencia
                    WHERE aco_codigo IN %s
                """, [tuple(user_aco_codigos)])
            else:
                cursor.execute("""
                    SELECT age_codigo, age_descricao
                    FROM agencia
                """)

        resultados = cursor.fetchall()

    # Formatar os resultados como uma lista de dicionários
    valores = [
        {'age_codigo': row[0], 'age_descricao': row[1]}
        for row in resultados
    ]


    return Response({'valores': valores})



def create_excel_byfilter(request) -> Response:
    user_id = request.user.id
    data_consulta = request.GET.get('dataInicio')
    data_consulta_final = request.GET.get('dataFim')
    unidade_selecionada = request.GET.get('unidade')
    areas_selecionadas = request.GET.getlist('areaComercial[]')
    agencia_selecionada = request.GET.getlist('agencias[]')
    vendedor_selecionada = request.GET.getlist('vendedor[]')

    # Consultando as áreas do usuário
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ac.aco_codigo
            FROM usuariocomercial ac
            INNER JOIN auth_user uac ON uac.id = ac.usr_codigo
            WHERE uac.id = %s
        """, [user_id])
        user_areas = [row[0] for row in cursor.fetchall()]

    # Validando as datas
    try:
        data_consulta_dt = datetime.strptime(data_consulta, "%Y-%m-%d")
        data_consulta_final_dt = datetime.strptime(data_consulta_final, "%Y-%m-%d")
    except ValueError:
        return Response(
            {"message": "Datas inválidas. Use o formato YYYY-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST
        )

    query = """
        SELECT fim_tipo, tur_numerovenda, tur_codigo, fim_valorliquido, fim_data, fim_markup, fim_valorinc, fim_valorincajustado, aco_descricao, age_descricao, ven_descricao, fat_valorvendabruta
        FROM faturamentosimplificado 
        WHERE fim_data BETWEEN %s AND %s 
    """
    params = [data_consulta_dt, data_consulta_final_dt]

    # Adicionando filtros conforme as seleções
    if user_areas:
        query += " AND aco_codigo IN %s"
        params.append(tuple(user_areas))

    if unidade_selecionada:
        query += " AND loj_codigo = %s"
        params.append(unidade_selecionada)

    if areas_selecionadas:
        query += " AND aco_codigo IN %s"
        params.append(tuple(areas_selecionadas))

    if agencia_selecionada:
        query += " AND age_codigo IN %s"
        params.append(tuple(agencia_selecionada))

    if vendedor_selecionada:
        query += " AND ven_codigo IN %s"
        params.append(tuple(vendedor_selecionada))

    query += " ORDER BY fim_data"

    # Executando a consulta e obtendo os resultados
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        resultados = cursor.fetchall()

    # Formatando os resultados com o serializer
    resultados_formatados = [RelatorioSerializer({
        'fim_tipo': resultado[0],
        'tur_numerovenda': resultado[1],
        'tur_codigo': resultado[2],
        'fim_valorliquido': resultado[3],
        'fim_data': resultado[4],
        'fim_markup': resultado[5],
        'fim_valorinc': resultado[6],
        'fim_valorincajustado': resultado[7],
        'aco_descricao': resultado[8],
        'age_descricao': resultado[9],
        'ven_descricao': resultado[10],
        'fat_valorvendabruta': resultado[11] if resultado[11] is not None else 0,
    }).data for resultado in resultados]

    # Chamar a função para processar os dados e gerar o Excel
    excel_data = process_data_chunk(resultados_formatados)

    # Calcular os totais após os filtros aplicados
    soma_totais = {
        'total_valorinc': sum(item['fim_valorinc'] for item in resultados_formatados if item['fim_valorinc'] is not None),
        'total_valorincajustado': sum(item['fim_valorincajustado'] for item in resultados_formatados if item['fim_valorincajustado'] is not None),
        'total_valorliquido': sum(item['fim_valorliquido'] for item in resultados_formatados if item['fim_valorliquido'] is not None),
    }

    # Retornar o Excel como resposta
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio.xlsx"'
    return response




def process_data_chunk(data_chunk):
    wb = Workbook()
    ws = wb.active
    headers = [
        'Tipo',
        'Núm. Venda',
        'Num. Pct',
        'Valor Liquido Venda',
        'Data',
        'MarkUp',
        'Valor Inc',
        'Inc Ajustado',
        'Area Comercial',
        'Agencia',
        'Vendedor',
        'Valor Venda Bruta'
    ]
    ws.append(headers)

    for relatorio in data_chunk:        
        # Adiciona os valores
        row = [
            relatorio['fim_tipo'],
            relatorio['tur_numerovenda'],
            relatorio['tur_codigo'],
            locale.currency(relatorio['fim_valorliquido'], grouping=True),
            relatorio['fim_data'],
            round(relatorio['fim_markup'], 4),
            locale.currency(relatorio['fim_valorinc'], grouping=True),
            locale.currency(relatorio['fim_valorincajustado'], grouping=True),
            relatorio['aco_descricao'],
            relatorio['age_descricao'],
            relatorio['ven_descricao'],
            locale.currency(relatorio['fat_valorvendabruta'], grouping=True),  # Aqui usamos o valor tratado
        ]
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()




def format_currency(value):
    # Verifica se o valor é None ou se não é um número válido
    if value is None:
        return "R$ 0,00"  # Ou qualquer outro valor default
    try:
        return locale.currency(value, grouping=True)
    except ValueError:
        return "R$ 0,00"  # Valor default em caso de erro

def list_all_area(request) -> Response:
    areas = AreaComercial.objects.all()
    serializer = AreaComercialSerializer(areas, many=True)  # Serialize os dados
    return Response(serializer.data)

def list_all_areas(request, unidade_id=None):
    user_id = request.user.id

    with connection.cursor() as cursor:
        if unidade_id:
            cursor.execute("""
                SELECT lc.aco_codigo, a.aco_descricao
                FROM lojacomercial lc
                INNER JOIN areacomercial a ON lc.aco_codigo = a.aco_codigo
                WHERE lc.loj_codigo = %s ORDER BY aco_descricao
            """, [unidade_id])
        else:
            cursor.execute("""
                SELECT a.aco_codigo, a.aco_descricao
                FROM areacomercial a ORDER BY aco_descricao
            """)

        resultados = cursor.fetchall()

    associacoes = [{'aco_codigo': row[0], 'aco_descricao': row[1]} for row in resultados]
    return Response({'associacoes': associacoes})




def obter_dados_unidade(request):
    # Parâmetros de filtro
    loj_codigo_param = request.GET.get('loj_codigo')  # Unidade
    date_start = request.GET.get('date_start')  # Data inicial
    date_end = request.GET.get('date_end')      # Data final

    # Consulta inicial
    queryset = Relatorio.objects.all()

    # Filtrar por intervalo de datas
    if date_start:
        start_date = parse_date(date_start)
        if start_date:
            queryset = queryset.filter(fim_data__gte=start_date)

    if date_end:
        end_date = parse_date(date_end)
        if end_date:
            queryset = queryset.filter(fim_data__lte=end_date)

    # Filtrar por unidade
    if loj_codigo_param and loj_codigo_param != "todos":
        queryset = queryset.filter(loj_codigo=loj_codigo_param)

    # Subquery para buscar o loj_descricao na tabela Loja
    loj_descricao_subquery = Loja.objects.filter(loj_codigo=OuterRef('loj_codigo')).values('loj_descricao')[:1]

    # Agrupar e somar os valores, adicionando a subquery como anotação
    resultados = (
        queryset
        .annotate(loj_descricao=Subquery(loj_descricao_subquery))
        .values('loj_descricao')
        .annotate(soma_valor=Round(Sum('fim_valorliquido'), 2))
        .order_by('-soma_valor')[:10]
    )

    # Formatar os dados para o gráfico
    dados_grafico = {
        'labels': [item['loj_descricao'] for item in resultados],
        'data': [item['soma_valor'] for item in resultados],
    }

    return Response(dados_grafico)


def obter_dados_agencia(request):
    # Parâmetros de filtro
    age_codigo_param = request.GET.getlist('age_codigo[]')  # Agência como lista
    date_start = request.GET.get('date_start')  # Data inicial
    date_end = request.GET.get('date_end')      # Data final

    # Consulta inicial
    queryset = Relatorio.objects.all()

    # Filtrar por intervalo de datas
    if date_start:
        start_date = parse_date(date_start)
        if start_date:
            queryset = queryset.filter(fim_data__gte=start_date)

    if date_end:
        end_date = parse_date(date_end)
        if end_date:
            queryset = queryset.filter(fim_data__lte=end_date)

    # Filtrar por agência
    if age_codigo_param and "todos" not in age_codigo_param:
        queryset = queryset.filter(age_codigo__in=age_codigo_param)

    # Agrupar e somar os valores
    resultados = (
        queryset.values('age_codigo', 'age_descricao')
        .annotate(soma_valor=Round(Sum('fim_valorliquido'), 2))
        .order_by('-soma_valor')[:10]
    )

    # Formatar os dados para o gráfico
    dados_grafico = {
        'labels': [item['age_descricao'] for item in resultados],
        'data': [item['soma_valor'] for item in resultados],
    }

    return Response(dados_grafico)


from django.http import JsonResponse

def obter_dados_area_comercial(request):
    if request.method == "POST":
        filters = request.data

        # Recupera os parâmetros da requisição
        date_start = filters.get('startDate')
        date_end = filters.get('endDate')
        aco_codigo_param = filters.get('areas', [])
        quantidade = filters.get('quantidade', 5)  # Padrão para 5 áreas, caso não seja passado

        queryset = Relatorio.objects.all()

        # Filtrar por intervalo de datas
        if date_start:
            try:
                start_date = datetime.strptime(date_start, "%Y-%m-%d")
                queryset = queryset.filter(fim_data__gte=start_date)
            except ValueError:
                return JsonResponse({"detail": "Formato de data inicial inválido"}, status=400)

        if date_end:
            try:
                end_date = datetime.strptime(date_end, "%Y-%m-%d")
                queryset = queryset.filter(fim_data__lte=end_date)
            except ValueError:
                return JsonResponse({"detail": "Formato de data final inválido"}, status=400)

        # Filtrar por Área Comercial
        if aco_codigo_param:
            queryset = queryset.filter(aco_codigo__in=aco_codigo_param)

        # Agrupar e somar os valores
        resultados = (
            queryset.values('aco_codigo', 'aco_descricao')
            .annotate(soma_valor=Sum('fim_valorliquido'))
            .order_by('-soma_valor')[:quantidade]  # Limitando pelos 'quantidade' de áreas
        )


        if not resultados:
            return JsonResponse({"labels": [], "data": []})

        dados_grafico = {
            'labels': [item['aco_descricao'] for item in resultados],
            'data': [item['soma_valor'] for item in resultados],
        }

        return JsonResponse(dados_grafico)

    return JsonResponse({"detail": "Método não permitido"}, status=405)



def exportar_dados_agencia_para_excel(request):
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # Definir local para Brasil (R$)

    # Parâmetros de filtro
    age_codigo_param = request.GET.getlist("age_codigo[]")  
    date_start = request.GET.get("date_start")  
    date_end = request.GET.get("date_end")
    unidade = request.GET.get("unidade")  
    area = request.GET.getlist("areas[]")  
    num_agencias = int(request.GET.get("num_agencias", 10))  

    # Se não tiver datas, usar mês atual
    if not date_start or not date_end:
        today = datetime.today()
        date_start = today.replace(day=1).strftime('%Y-%m-%d')
        date_end = today.replace(day=28).strftime('%Y-%m-%d')

    start_date = parse_date(date_start)
    end_date = parse_date(date_end)

    # Consulta inicial
    queryset = Relatorio.objects.all()

    # Aplicar filtros
    if start_date:
        queryset = queryset.filter(fim_data__gte=start_date)

    if end_date:
        queryset = queryset.filter(fim_data__lte=end_date)

    if age_codigo_param and "todos" not in age_codigo_param:
        queryset = queryset.filter(age_codigo__in=age_codigo_param)

    if unidade:
        queryset = queryset.filter(loj_codigo=unidade)

    if area:  # Filtro corrigido para área comercial
        queryset = queryset.filter(aco_codigo__in=area)

    # Log dos filtros aplicados
    print(f"Filtro unidade: {unidade}, Filtro área: {area}")

    # Agrupar e somar valores
    resultados = (
        queryset.values('age_codigo', 'age_descricao')
        .annotate(total_vendas=Round(Sum('fim_valorliquido'), 2))
        .order_by('-total_vendas')[:num_agencias]
    )

    # Criar planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Agências"

    # Cabeçalhos
    headers = ["Código Agência", "Agência", "Total de Vendas"]
    ws.append(headers)

    # Estilizar cabeçalho
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Adicionar dados
    total_geral = 0  
    current_row = 2  

    if resultados.exists():
        for item in resultados:
            agency_code = item['age_codigo']
            agency_desc = item['age_descricao']
            agency_total = item['total_vendas']

            formatted_agency_total = locale.currency(agency_total, grouping=True, symbol=True)

            ws.append([agency_code, agency_desc, formatted_agency_total])
            total_geral += agency_total
            current_row += 1  

        # Adicionar linha do total geral
        ws.append(["", "Total Geral:", locale.currency(total_geral, grouping=True, symbol=True)])
        ws[f"B{current_row}"].font = Font(bold=True)
    else:
        ws.append(["Nenhum dado encontrado para os filtros aplicados."])

    # Retornar arquivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="relatorio_agencias_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'

    wb.save(response)
    return response

import csv
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def exportar_dados_loja_para_excel(request):
    # Configura o local para o Brasil (R$)
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # Ajuste conforme necessário

    # Parâmetros de filtro
    unidade_codigo_param = request.GET.getlist("unidade_codigo[]")  
    date_start = request.GET.get("date_start")  
    date_end = request.GET.get("date_end")      
    num_unidades = int(request.GET.get("num_unidades", 10))  # Default para 10

    # Se não foi passado o filtro de data, usar o mês atual
    if not date_start or not date_end:
        today = datetime.today()
        date_start = today.replace(day=1).strftime('%Y-%m-%d')  # Primeiro dia do mês atual
        date_end = today.replace(day=28).strftime('%Y-%m-%d')  # Último dia do mês atual (geralmente 28)

    # Converter as strings de data em objetos de data
    start_date = parse_date(date_start)
    end_date = parse_date(date_end)

    # Consulta inicial
    queryset = Relatorio.objects.all()

    # Aplicar filtros de data
    if start_date:
        queryset = queryset.filter(fim_data__gte=start_date)

    if end_date:
        queryset = queryset.filter(fim_data__lte=end_date)

    # Filtrar por unidade
    if unidade_codigo_param and "todos" not in unidade_codigo_param:
        queryset = queryset.filter(loj_codigo__in=unidade_codigo_param)  # Filtrando pela unidade (loj_codigo)

    # Ordenar pelas unidades com maior valor total de vendas e limitar o número de resultados
    queryset = queryset.values('loj_codigo') \
                       .annotate(total_vendas=Sum('fim_valorliquido')) \
                       .order_by('-total_vendas')[:num_unidades]  # Limitar pelas maiores vendas

    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Lojas"

    # Criar cabeçalho
    headers = [
        "Código Loja", "Loja", "Total Valor Líquido"
    ]
    ws.append(headers)

    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Cinza claro
    header_font = Font(bold=True, color="000000")  # Negrito e preto
    header_border = Border(
        left=Side(style="thin", color="E1E1E1"),  # Cor mais suave para as bordas
        right=Side(style="thin", color="E1E1E1"),
        top=Side(style="thin", color="E1E1E1"),
        bottom=Side(style="thin", color="E1E1E1")
    )

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border

    # Configuração das colunas: Largura, Alinhamento e Estilo
    colunas_config = {
        "A": 12,  # Código Loja
        "B": 25,  # Nome da Loja
        "C": 15,  # Total Valor Líquido
    }

    # Aplicar largura e alinhamento às colunas
    for col, width in colunas_config.items():
        ws.column_dimensions[col].width = width
        for cell in ws[col]:  # Itera sobre todas as células da coluna
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adicionar dados ao Excel
    total_geral = 0  # Inicializa a variável para o total geral
    last_loja = None  # Controlar a mudança de loja para mostrar os registros dela
    current_row = 2  # A partir da linha 2 para os dados

    for item in queryset:
        # Acessando os dados como dicionário
        loj_codigo = item['loj_codigo']  # Código da loja

        # Obter a descrição da loja a partir do modelo Loja
        try:
            loja = Loja.objects.get(loj_codigo=loj_codigo)
            loja_desc = loja.loj_descricao  # Ajuste conforme o nome da descrição da loja
        except Loja.DoesNotExist:
            loja_desc = "Loja Desconhecida"

        if last_loja != loj_codigo:  # Se a loja mudar, adicionar uma linha para a mudança
            if last_loja is not None:  # Adicionar uma linha em branco entre lojas
                ws.append(["", "", ""])  # Linha em branco

            last_loja = loj_codigo  # Atualiza a loja atual

        # Somar os valores para o total da loja
        loja_total = item['total_vendas']  # A soma já foi calculada na consulta

        # Adicionar apenas o total da unidade (sem registros individuais)
        ws.append([
            loj_codigo,
            loja_desc,
            locale.currency(loja_total, grouping=True, symbol=True)
        ])

        total_geral += loja_total  # Somar o total da loja ao total geral
        current_row += 1  # Mover para a próxima linha após adicionar o total da unidade

    # Adicionar a linha do total geral no final
    ws.append(["", "Total Geral:",locale.currency(total_geral, grouping=True, symbol=True), ""])

    # Criar a resposta HTTP com o arquivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="relatorio_lojas_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'

    # Salvar o arquivo na resposta
    wb.save(response)

    return response


from collections import defaultdict
from openpyxl.styles import Font, Alignment

def exportar_area_comercial_para_excel(request):
    filters = request.GET

    date_start = filters.get("date_start")
    date_end = filters.get("date_end")
    aco_codigo_param = filters.getlist('areas[]')  # Usando getlist para pegar o array com o nome correto
    print("Áreas recebidas:", aco_codigo_param)  # Para depurar se o filtro foi recebido corretamente
    
    # Certificando-se de que a lista de áreas seja uma lista de inteiros válidos
    aco_codigo_param = [int(area) for area in aco_codigo_param if area.isdigit()]
    quantidade = int(filters.get("quantidade", 5))  # Quantidade padrão: 5

    # Validação de datas
    if not date_start or not date_end:
        today = datetime.today()
        date_start = today.replace(day=1).strftime("%Y-%m-%d")
        date_end = today.replace(day=28).strftime("%Y-%m-%d")

    start_date = datetime.strptime(date_start, "%Y-%m-%d")
    end_date = datetime.strptime(date_end, "%Y-%m-%d")

    queryset = Relatorio.objects.all()

    # Filtrando pela data
    queryset = queryset.filter(fim_data__range=[start_date, end_date])

    # Filtrando pelas áreas comerciais SE as áreas forem passadas
    if aco_codigo_param:
        queryset = queryset.filter(aco_codigo__in=aco_codigo_param)  # Filtrando pelas áreas selecionadas
        print(f"Filtrando pelas áreas comerciais: {aco_codigo_param}")
    else:
        print("Nenhuma área comercial foi selecionada. Aplicando filtro por quantidade.")

    # Agrupar por área comercial e calcular total
    area_comercial_totals = (
        queryset.values("aco_codigo", "aco_descricao")
        .annotate(soma_valor=Sum("fim_valorliquido"))
        .order_by("-soma_valor")  # Ordenando do maior para o menor valor
    )

    # Aplicando o limite de quantidade
    area_comercial_totals = list(area_comercial_totals)[:quantidade]

    # Calculando o total geral
    total_geral = sum(area["soma_valor"] for area in area_comercial_totals)

    # Criando o arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Área Comercial"

    headers = ["Código Área Comercial", "Área Comercial", "Valor Total"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)

    # Preenchendo as linhas com dados
    for area in area_comercial_totals:
        ws.append([ 
            area["aco_codigo"],
            area["aco_descricao"],
            locale.currency(area["soma_valor"], grouping=True, symbol=True)  # Formatação correta
        ])

    # Adicionando o total geral
    ws.append(["", "TOTAL GERAL", locale.currency(total_geral, grouping=True, symbol=True)])

    # Ajustando o tamanho das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Retornando o arquivo Excel como resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="relatorio_area_comercial_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
    wb.save(response)

    return response
